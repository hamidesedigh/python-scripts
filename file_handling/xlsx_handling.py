# -*- coding: utf-8 -*-
"""
@author: Hamideh
Created on Wed Oct  8 09:53:22 2025

Program Description:
-------------------------
Demonstration of Excel File Handling using openpyxl
--------------------------------------------------------------
Covers:
✔ Workbook & Worksheet creation
✔ Writing & Reading Excel files
✔ Multiple sheet management
✔ City summary statistics (count, total, average)
"""

from openpyxl import Workbook, load_workbook
from collections import defaultdict

# ============================================================
# --- Step 1: Create Excel Workbook ---
# ============================================================

excel_file = "people_data.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "People"

headers = ["Name", "Age", "City", "Income"]
data = [
    ["Hamideh", 39, "Berlin", 75000],
    ["Ali", 34, "Tehran", 56000],
    ["Sara", 29, "Paris", 61000],
    ["John", 45, "London", 82000],
    ["Mary", 31, "Berlin", 72000]
]

ws.append(headers)
for row in data:
    ws.append(row)

wb.save(excel_file)
print("✅ Excel file 'people_data.xlsx' created successfully.\n")

# ============================================================
# --- Step 2: Read Excel Data ---
# ============================================================

wb2 = load_workbook(excel_file)
ws2 = wb2.active

print("--- Reading Excel Data ---")
for row in ws2.iter_rows(values_only=True):
    print(row)

# ============================================================
# --- Step 3: Add City Summary Sheet ---
# ============================================================

print("\n📊 Creating City Summary...")

city_stats = defaultdict(lambda: {"Count": 0, "Total Income": 0, "Average Income": 0})

for name, age, city, income in data:
    city_stats[city]["Count"] += 1
    city_stats[city]["Total Income"] += income

for city, stats in city_stats.items():
    stats["Average Income"] = stats["Total Income"] / stats["Count"]

ws_summary = wb2.create_sheet("City_Summary")
ws_summary.append(["City", "Count", "Total Income (€)", "Average Income (€)"])

for city, stats in city_stats.items():
    ws_summary.append([
        city,
        stats["Count"],
        stats["Total Income"],
        round(stats["Average Income"], 2)
    ])

wb2.save("people_data_extended.xlsx")
print("✅ 'City_Summary' sheet added to 'people_data_extended.xlsx'.")

# ============================================================
# --- Quick Summary ---
# ============================================================
"""
Quick Summary:
--------------
✔ Workbook() → Creates new Excel workbook
✔ .active / .create_sheet() → Manage sheets
✔ .append() → Add rows
✔ iter_rows() → Read Excel data easily
✔ defaultdict() → Useful for grouped calculations
✔ Created summary sheet with statistics by city
"""
